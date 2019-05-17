import re
from xml.dom.minidom import parse
import xml.dom.minidom

#open the xml file and create the DOM tree
DOMTree = xml.dom.minidom.parse("go_obo.xml")
collection = DOMTree.documentElement
terms = collection.getElementsByTagName("term")

genelists = [["id","name","defination","childnodes"],]
genelist = []


#function to count the childnodes
def Child(id, resultSet):
    for t in terms:
        parents = t.getElementsByTagName("is_a")
        geneid = t.getElementsByTagName('id')[0].childNodes[0].nodeValue
        for parent in parents:
            if parent.childNodes[0].data == id:
                resultSet.add(geneid)
                Child(geneid,resultSet)


    
        

for term in terms:
    defstr = term.getElementsByTagName('defstr')[0].childNodes[0].nodeValue
    id = term.getElementsByTagName('id')[0].childNodes[0].nodeValue
    name = term.getElementsByTagName('name')[0].childNodes[0].nodeValue

#select the "autophagosome" terms        
    if re.search("autophagosome",defstr):
        resultSet = set()
        #count the childnodes
        Child(id, resultSet)
        genelist = [id, name, defstr, len(resultSet)]
        genelists.append(genelist)




#Creat an excel for autophagosome
import openpyxl
 
 
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx success")

def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()
 
 
book_name_xlsx = 'autophagosome.xlsx'
 
sheet_name_xlsx = 'autophagosome'
 
value3 = genelists
 
 
write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)

 













    





    
    
